####################################################################################
# AUTHOR: LAURENZ MARI A. REPASO
# DATE: 01/28/2025

# To save time in the daily warehousing task and avoid errors
# made by copy and pasting data.
####################################################################################

import streamlit as st
import pandas as pd
from io import BytesIO
import re
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# st.image("C:/Users/FC/Downloads/Scripts/fclogo.jfif", width=100)
st.title("Daily Warehousing")
st.text("Exports the collected cheques in the format of Union Bank's spreadsheet for warehousing.\nby: Laurenz Repaso")

input_today_date = st.date_input("Enter today's date:")
input_collected_date = st.date_input("Enter date of cheque collection:")

uploaded_file = st.file_uploader("Upload file (includes Regular Loans, SPV2 Replacement and Restructuring):", type=["xlsx"])

if uploaded_file:
    reg_loans = pd.read_excel(uploaded_file, sheet_name="Regular Loans", dtype=str)
    spv2_rc = pd.read_excel(uploaded_file, sheet_name="SPV2 Replacement Cheques", dtype=str)
    spv2_rst = pd.read_excel(uploaded_file, sheet_name="SPV2 Restructuring", dtype=str)

    # spv1_rst = pd.read_excel(uploaded_file, sheet_name="SPV1 Restructuring", dtype=str)
    # spv1_rc = pd.read_excel(uploaded_file, sheet_name="SPV1 Replacement Cheques", dtype=str)

    cvr = pd.read_excel(uploaded_file, sheet_name="Cheque Verification Request", dtype=str)

    # df = pd.concat([reg_loans,spv2_rc, spv1_rc, spv1_rst, spv2_rst, cvr], ignore_index=True)
    df = pd.concat([reg_loans,spv2_rc, spv2_rst, cvr], ignore_index=True)

    st.success("File loaded successfully!")

    if input_collected_date and input_today_date:
        # convert 'collected_date' to datetime for filtering
        df["collected_date"] = pd.to_datetime(df["collected_date"], errors="coerce")
        df["cheque_date"] = pd.to_datetime(df["cheque_date"], errors="coerce")

        # filter data where 'collected_date' matches the input date
        filtered_data = df[df["collected_date"] == pd.to_datetime(input_collected_date)]

        # separates warehousing and manual deposit cheques
        warehouse = filtered_data[(filtered_data["cheque_date"] - pd.to_datetime(input_today_date)).dt.days >= 7]
        manual_deposit = filtered_data[(filtered_data["cheque_date"] - pd.to_datetime(input_today_date)).dt.days < 7]

        # sort cheques for easy cross checking
        warehouse = warehouse.sort_values(by=["collected_date", "bank_name"], ascending=[True, True])
        
        # add column LOCAL
        warehouse["blank_column"] = "LOCAL"
        manual_deposit["blank_column"] = "LOCAL"
        

        # remove special characters
        warehouse["company_name"] = warehouse["company_name"].apply(
            lambda x: re.sub(r"[^a-zA-Z0-9\s]", "", x)
            if isinstance(x, str)
            else x)
        manual_deposit["company_name"] = manual_deposit["company_name"].apply(
            lambda x: re.sub(r"[^a-zA-Z0-9\s]", "", x)
            if isinstance(x, str)
            else x)

        # fixed number of digits with leading zeroes
        warehouse["brstn_code"] = warehouse["brstn_code"].apply(
            lambda x:f"{int(x):09}"
            if str(x).isdigit()
            else x)
        # fixed number of digits with leading zeroes
        manual_deposit["brstn_code"] = manual_deposit["brstn_code"].apply(
            lambda x:f"{int(x):09}"
            if str(x).isdigit()
            else x)
        
        # rearrange columns
        warehouse = warehouse[
            [
                "cheque_amount",             # Amount column
                "brstn_code",                # ID (as string) 
                "bank_account_number",       # ID (as string)
                "cheque_identifier",         # ID (as string)
                "cheque_date",               # Date
                "blank_column",              # LOCAL
                "company_name",              # String
                "lead_id",                   # ID (as string)
            ]
        ]

                # rearrange columns
        manual_deposit = manual_deposit[
            [
                "cheque_amount",             # Amount column
                "brstn_code",                # ID (as string) 
                "bank_account_number",       # ID (as string)
                "cheque_identifier",         # ID (as string)
                "cheque_date",               # Date
                "blank_column",              # LOCAL
                "company_name",              # String
                "lead_id",                   # ID (as string)
            ]
        ]

        # date format for UB hub
        warehouse["cheque_date"] = pd.to_datetime(warehouse["cheque_date"], errors="coerce").dt.strftime("%m/%d/%Y")
        manual_deposit["cheque_date"] = pd.to_datetime(manual_deposit["cheque_date"], errors="coerce").dt.strftime("%m/%d/%Y")

        # turn to string to avoid scientific notations
        id_columns = ["brstn_code", "bank_account_number", "cheque_identifier", "lead_id"]
        for col in id_columns:
            warehouse[col] = warehouse[col].astype(str)
            manual_deposit[col] = manual_deposit[col].astype(str)

        # remove leading zeroes 
        warehouse["cheque_identifier"] = warehouse["cheque_identifier"].apply(
            lambda x: str(int(x))
            if str(x).isdigit()
            else x
            )
        manual_deposit["cheque_identifier"] = manual_deposit["cheque_identifier"].apply(
            lambda x: str(int(x))
            if str(x).isdigit()
            else x
            )

        def export_to_excel(warehouse, manual_deposit):
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                warehouse.to_excel(writer, index=False, sheet_name="For warehousing")
                manual_deposit.to_excel(writer, index=False, sheet_name="For manual deposit")
            output.seek(0)
            return output.getvalue()


        # for cross checking the number of cheques and cheque amount
        cheque_count_warehouse = len(warehouse)
        total_cheque_amount_w = warehouse["cheque_amount"].astype(float).sum()

        cheque_count_md = len(manual_deposit)
        total_cheque_amount_md = manual_deposit["cheque_amount"].astype(float).sum()

        # warehousing
        st.subheader("Warehouse")
        st.write(f"**Number of cheques:** {cheque_count_warehouse}")
        st.write(f"**Total cheque amount:** ₱{total_cheque_amount_w:,.2f}")

        # manual deposit
        st.subheader("Manual Deposit")
        st.write(f"**Number of cheques:** {cheque_count_md}")
        st.write(f"**Total cheque Amount:** ₱{total_cheque_amount_md:,.2f}")

        filename = f"cheques@{input_collected_date.strftime('%Y/%m/%d')}.xlsx"

        # buttons
        col1, col2 = st.columns([1,1])
        with col1:
            st.download_button(
                label="Download Excel File",
                data=export_to_excel(warehouse, manual_deposit),
                file_name=f"cheques_summary@{input_collected_date.strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        with col2:
            st.link_button("UnionBank Hub", "https://hub.unionbankph.com/b2bhub/")


